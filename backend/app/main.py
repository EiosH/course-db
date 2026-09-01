import argparse
import json
import re
import time
from datetime import datetime
from types import SimpleNamespace

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from qdrant_client import QdrantClient
from qdrant_client.models import (
    Distance,
    Document,
    FieldCondition,
    Filter,
    MatchValue,
    Modifier,
    PayloadSchemaType,
    PointStruct,
    PointVectors,
    Range,
    SparseVectorParams,
    VectorParams,
)

OLLAMA_URL = "http://127.0.0.1:11434"
OLLAMA_MODEL = "qwen3.8:27b"
EMBED_MODEL = "bge-m3"
BM25_MODEL = "Qdrant/bm25"
RERANK_MODEL = "Qwen/Qwen3-Reranker-0.6B"
OLLAMA_CHAT_TIMEOUT = 600
OLLAMA_CHAT_RETRIES = 3

# mock 预过滤
COURSE_ID = "CSC447"
QUARTER = "2026-Spring"
LECTURER = "Eric J. Fredericks"
TIMESTAMP = "01:22:09"  # mock 当前播放位置，仅时间类问题启用
TIME_WINDOW_SEC = 120  # 时间戳约束：±2 分钟，优先当前画面/台词
BASE_MUST = [
    FieldCondition(key="course_id", match=MatchValue(value=COURSE_ID)),
    FieldCondition(key="quarter", match=MatchValue(value=QUARTER)),
    FieldCondition(key="lecturer", match=MatchValue(value=LECTURER)),
]

# 每路召回候选数 → rerank 后保留
DENSE_LIMIT = 10
BM25_LIMIT = 10
TIME_NEAR_TOP_K = 4  # 时间类问题：最终保留条数（screen_shot / transcript 各半）
RERANK_TOP_K = 4
RERANK_MIN_SCORE = 0.0  # Qwen3-Reranker: yes/no logit diff，>0 表示相关

REWRITE_SYSTEM = f"""You are a search query rewrite assistant for lecture retrieval.

Current playback timestamp (mock): {TIMESTAMP}

Output JSON only:
{{
  "rewritten_query": "...",
  "hard_constraints": []
}}

Rules:
- rewritten_query: informative search string for lecture notes, screenshot OCR, and transcripts (not a tiny 3-word stub)
- Keep concrete labels and domain terms. Do NOT include course codes, instructor names, or quarter — those are already filtered
- If the user refers to a numbered quiz/homework item (e.g. "Question 9", "Problem 3", "Q9", "exercise 2"):
  * KEEP the exact label in rewritten_query (e.g. "Question 9")
  * Expand with retrieval intents: full question stem, options/choices, correct answer, instructor explanation
  * Example shape: "Question 9 quiz stem options answer explanation"
  * Do NOT invent the actual question text — you do not know it yet; only keep the label + intents
- For other questions: expand with synonyms and technical terms that likely appear on slides/transcripts
- hard_constraints: pre-filters for retrieval
- ONLY add {{"field":"timestamp","operator":"range","value":"{TIMESTAMP}"}} when the user explicitly asks about what is being discussed RIGHT NOW at the current moment (e.g. "现在在讲什么", "what are we covering now", "what does this mean now")
- When adding a timestamp constraint, rewritten_query can be a short placeholder (e.g. "current slide and transcript"); retrieval will use the timestamp, not keywords
- Do NOT add timestamp for general knowledge questions, definitions, or past/future topics
- If no timestamp constraint is needed, hard_constraints must be []"""

# 多词定位短语（任意 "Label + 数字"）：收成单个 BM25 token，查询侧同步扩词。
# token = "ph" + 去空白标点(label+num)，由规则动态生成（不是写死某个题号）。
ANCHOR_PHRASE_RE = re.compile(
    r"(?i)\b([a-z]+)(?:\s*#\s*|\s+)(\d+)\b"
)
ANCHOR_STOPWORDS = frozenset(
    {
        "a",
        "an",
        "the",
        "and",
        "or",
        "of",
        "to",
        "in",
        "on",
        "for",
        "is",
        "are",
        "was",
        "be",
        "by",
        "at",
        "as",
        "it",
        "this",
        "that",
        "with",
        "from",
        "not",
        "do",
        "does",
        "did",
        "if",
        "so",
        "no",
        "yes",
        "my",
        "your",
        "our",
        "their",
        "line",
        "row",
        "col",
        "item",  # 太泛；真正的 Item 12 若需要可再放回 HINT
    }
)
# 常见定位词；其它非停用词 Label+数字也可（如 Lab 3 / Assignment 2）
ANCHOR_HINT_WORDS = frozenset(
    {
        "question",
        "problem",
        "exercise",
        "quiz",
        "homework",
        "hw",
        "slide",
        "page",
        "part",
        "section",
        "chapter",
        "lecture",
        "unit",
        "module",
        "topic",
        "task",
        "lab",
        "assignment",
        "homework",
        "q",
    }
)
ANCHOR_LIMIT = 8
ANCHOR_EXPAND_CHARS = 900

ANSWER_SYSTEM = """You are a friendly course assistant sitting next to the student during lecture.

Tone:
- Warm, brief, human — like a helpful classmate, not a search system.
- NEVER use machine phrases such as:
  "Based on the retrieved lecture materials",
  "I don't know based on the retrieved lecture materials",
  "According to the provided context",
  "The retrieved materials do not contain…".
- Answer directly. Prefer "The slide shows…" / "The instructor said…".

Rules:
1. Use ONLY facts from the lecture content below. No outside knowledge.
2. Be concise; quote a short phrase when helpful.
3. If the content includes a quiz/homework item the student asked about, briefly restate that item (stem / key options) then explain the answer using the materials.
4. If the content is empty or doesn't cover the question, apologize like a person, e.g.:
   "Hmm, I couldn't find that in this lecture — sorry, I'm not sure."
   "I looked through the notes but nothing on that popped up. Want to try rephrasing?"
   Do NOT use stiff template refusals."""

NOW_ANSWER_SYSTEM = """You are a friendly course assistant. The student is asking what the lecture is covering RIGHT NOW (current slide and/or transcript near the playback time).

Tone:
- Warm and human — like you're watching the lecture with them.
- NEVER use "Based on the retrieved lecture materials", "I don't know based on the retrieved…", or similar.
- Prefer "Right now the slide is about…" / "The instructor is explaining…".

Rules:
1. The content below IS what's happening now. Explain it even if the question is vague.
2. Use ONLY those facts. No outside knowledge.
3. Be concise; mention slide title, code, or a key spoken line when helpful.
4. If content is empty, say something like:
   "I don't have anything from this moment in the lecture — sorry!"
"""

# 无检索结果时不交给模型套模板，直接用人话回复
NO_HIT_REPLY = (
    "Hmm, I couldn't find anything useful in the lecture notes for that — "
    "sorry, I'm not sure. Want to try asking another way?"
)
NO_HIT_NOW_REPLY = (
    "I don't have anything from this moment in the lecture — sorry! "
    "Maybe scrub a bit or ask about a specific term on the slide."
)

# 模型仍可能吐出的生硬拒答 → 替换成助手语气
STIFF_REFUSAL_RE = re.compile(
    r"(?is)^\s*(?:based on (?:the )?(?:retrieved )?lecture materials[,.]?\s*)?"
    r"i don'?t know(?: based on the retrieved lecture materials)?\.?\s*$"
)
STIFF_REFUSAL_REPLY = (
    "Hmm, I couldn't find that in this lecture — sorry, I'm not sure. "
    "Want to try rephrasing?"
)

# transcript：合并相邻字幕，约 400 token；静音超过 15s 切新块
TRANSCRIPT_MAX_CHARS = 1600
TRANSCRIPT_MAX_GAP_SEC = 15
CUE_RE = re.compile(
    r"(\d{2}:\d{2}:\d{2}(?:\.\d+)?)\s*-->\s*(\d{2}:\d{2}:\d{2}(?:\.\d+)?)\s*\n(.*?)(?=\n\d{2}:\d{2}:\d{2}|\Z)",
    re.S,
)
SCREEN_LABEL_RE = re.compile(
    r"(?im)^(?:Type|Top|Left|Right|Bottom|Content|Title|Bullet Points|Notes|Question)\s*:\s*"
)
SCREEN_BOILERPLATE_RE = re.compile(
    r"(?im)(?:reed\.cs\.depaul\.edu[^\n]*|to exit full screen[^\n]*|press Esc[^\n]*)\n?"
)

EMBED_BATCH = 32
_reranker = None


def get_reranker():
    global _reranker
    if _reranker is None:
        from sentence_transformers import CrossEncoder

        print(f"loading rerank model {RERANK_MODEL}...")
        _reranker = CrossEncoder(RERANK_MODEL, trust_remote_code=True)
    return _reranker


def embed(texts):
    out = []
    total = len(texts)
    for i in range(0, total, EMBED_BATCH):
        batch = texts[i : i + EMBED_BATCH]
        r = requests.post(
            f"{OLLAMA_URL}/api/embed",
            json={"model": EMBED_MODEL, "input": batch},
            timeout=300,
        )
        if not r.ok:
            raise RuntimeError(f"Ollama embed {r.status_code}: {r.text}")
        out.extend(r.json()["embeddings"])
        done = min(i + EMBED_BATCH, total)
        print(f"embed {done}/{total}")
    return out


def ollama_chat(messages, *, format=None, timeout=OLLAMA_CHAT_TIMEOUT):
    """Call Ollama /api/chat with retries; think=False avoids qwen thinking overhead."""
    payload = {
        "model": OLLAMA_MODEL,
        "messages": messages,
        "stream": False,
        "think": False,
    }
    if format is not None:
        payload["format"] = format

    last_err = None
    for attempt in range(1, OLLAMA_CHAT_RETRIES + 1):
        try:
            r = requests.post(
                f"{OLLAMA_URL}/api/chat",
                json=payload,
                timeout=timeout,
            )
            r.raise_for_status()
            return r.json()["message"]["content"].strip()
        except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectionError) as e:
            last_err = e
            if attempt < OLLAMA_CHAT_RETRIES:
                wait = 5 * attempt
                print(
                    f"ollama chat timeout/error (attempt {attempt}/{OLLAMA_CHAT_RETRIES}), "
                    f"retry in {wait}s..."
                )
                time.sleep(wait)
    raise last_err


def rewrite(query: str) -> dict:
    raw = ollama_chat(
        [
            {"role": "system", "content": REWRITE_SYSTEM},
            {"role": "user", "content": query},
        ],
        format="json",
    )
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
    data = json.loads(raw)
    return {
        "rewritten_query": data["rewritten_query"],
        "hard_constraints": data.get("hard_constraints", []),
    }


def extract_anchor_phrases(*texts: str) -> list[str]:
    """Pull 'Label + number' locators from user/rewrite/chunk text (course-agnostic)."""
    found, seen = [], set()
    for text in texts:
        if not text:
            continue
        for m in ANCHOR_PHRASE_RE.finditer(text):
            label, num = m.group(1), m.group(2)
            phrase = normalize_anchor_phrase(f"{label} {num}")
            if not is_useful_anchor_phrase(phrase):
                continue
            key = phrase.lower()
            if key in seen:
                continue
            seen.add(key)
            found.append(phrase)
    return found


def is_useful_anchor_phrase(phrase: str) -> bool:
    """Keep real locators; drop stopword noise like 'the 9'."""
    m = re.match(r"(?i)^([a-z]+)\s+(\d+)$", phrase.strip())
    if not m:
        return False
    word = m.group(1).lower()
    if word in ANCHOR_HINT_WORDS:
        return True
    return word not in ANCHOR_STOPWORDS


def normalize_anchor_phrase(phrase: str) -> str:
    """Collapse spaces; Q9 → Question 9; capitalize label lightly."""
    phrase = re.sub(r"\s+", " ", phrase).strip()
    m = re.fullmatch(r"(?i)q\s*#?\s*(\d+)", phrase)
    if m:
        return f"Question {m.group(1)}"
    m = re.match(r"(?i)^([a-z]+)\s*#?\s*(\d+)\s*$", phrase)
    if not m:
        return phrase
    word, num = m.group(1).lower(), m.group(2)
    aliases = {"hw": "Homework", "q": "Question"}
    return f"{aliases.get(word, word.capitalize())} {num}"


def phrase_token(phrase: str) -> str:
    """
    Any Label+number locator → one BM25 term (computed, never hardcoded).
    formula: "ph" + alnum(normalize(phrase))
    """
    norm = normalize_anchor_phrase(phrase).lower()
    return "ph" + re.sub(r"[^a-z0-9]+", "", norm)

def phrases_in_text(text: str) -> list[str]:
    """Locator phrases appearing inside a chunk (for BM25 index enrichment)."""
    return extract_anchor_phrases(text)


def bm25_index_text(text: str) -> str:
    """Original text + dynamically derived whole-phrase tokens for sparse BM25."""
    tokens = [phrase_token(p) for p in phrases_in_text(text)]
    if not tokens:
        return text
    seen, uniq = set(), []
    for t in tokens:
        if t not in seen:
            seen.add(t)
            uniq.append(t)
    return text + "\n" + " ".join(uniq)


def bm25_phrase_query(*texts: str) -> str:
    """Query-side expansion: emit glued tokens for whatever locators appear."""
    tokens = [phrase_token(p) for p in extract_anchor_phrases(*texts)]
    seen, uniq = set(), []
    for t in tokens:
        if t not in seen:
            seen.add(t)
            uniq.append(t)
    return " ".join(uniq)


def search_phrase_bm25(client, phrase: str, doc_type: str, constraints, limit=ANCHOR_LIMIT):
    """BM25 with the glued phrase token (needs ingest/backfill that wrote those tokens)."""
    token = phrase_token(phrase)
    hits = search_bm25(client, token, doc_type, constraints, limit=limit)
    kept = [
        h for h in hits if text_has_phrase(h.payload.get("text", ""), phrase)
    ]
    return kept or hits


def text_has_phrase(text: str, phrase: str) -> bool:
    """Check payload literally contains the locator (generic; not quiz-specific)."""
    if not text:
        return False
    low = text.lower()
    canon = normalize_anchor_phrase(phrase)
    variants = [canon, phrase]
    # compact form: "Question 9" ↔ "Q9" only when first word starts with Q-word single letter alias
    m = re.match(r"(?i)^([a-z]+)\s+(\d+)$", canon)
    if m:
        word, num = m.group(1), m.group(2)
        variants.append(f"{word} {num}")
        variants.append(f"{word}{num}")
        if word.lower() == "question":
            variants.extend([f"Q{num}", f"Q {num}"])
        if word.lower() == "homework":
            variants.extend([f"HW{num}", f"HW {num}"])
    seen = set()
    for v in variants:
        k = v.lower()
        if k in seen:
            continue
        seen.add(k)
        parts = [re.escape(p) for p in re.split(r"\s+", k) if p]
        if not parts:
            continue
        # also allow zero-space compact match already in variants
        pat = r"\s*".join(parts)
        if re.search(rf"(?<![a-z0-9]){pat}(?![a-z0-9])", low):
            return True
    return False


def phrase_match_score(text: str, phrase: str) -> int:
    if not text_has_phrase(text, phrase):
        return 0
    score = 3
    low = text.lower()
    for cue in (
        "options",
        "correct",
        "points",
        "consider the",
        "which of the following",
        "answer",
    ):
        if cue in low:
            score += 1
    return score


def expand_query_with_anchors(client, query: str, rewritten_q: str, constraints):
    """
    Query-side phrase expansion (generic):
    - detect any Label+number locator in the user question
    - BM25-search its glued token (same token written at ingest)
    - append matching chunk text into the search query
    """
    phrases = extract_anchor_phrases(query, rewritten_q)
    if not phrases:
        return rewritten_q, []

    phrase_q = bm25_phrase_query(query, rewritten_q)
    print(f"phrase expand tokens: {phrase_q!r} from {phrases}")

    anchor_hits, snippets = [], []
    for phrase in phrases:
        raw_hits = merge_unique_hits(
            search_phrase_bm25(
                client, phrase, "screen_shot", constraints, limit=ANCHOR_LIMIT
            ),
            search_phrase_bm25(
                client,
                phrase,
                "transcript",
                constraints,
                limit=max(3, ANCHOR_LIMIT // 2),
            ),
        )
        ranked = sorted(
            raw_hits,
            key=lambda h: (
                -phrase_match_score(h.payload.get("text", ""), phrase),
                -(getattr(h, "score", None) or 0),
            ),
        )
        for h in ranked[:4]:
            if any(a.id == h.id for a in anchor_hits):
                continue
            anchor_hits.append(h)
            text = h.payload.get("text", "").strip()
            if text:
                snippets.append(text[:ANCHOR_EXPAND_CHARS])

    base = rewritten_q if not phrase_q else f"{rewritten_q}\n{phrase_q}"

    if not snippets:
        return base, []

    unique_snips = []
    for s in snippets:
        head = s[:160]
        if any(head in u or u[:160] in s for u in unique_snips):
            continue
        unique_snips.append(s)
        if len(unique_snips) >= 2:
            break

    expanded = (
        f"{base}\n"
        f"Related lecture excerpt for {' / '.join(phrases)}:\n"
        + "\n---\n".join(unique_snips)
    )
    print(
        f"anchor expand phrases={phrases} hits={len(anchor_hits)} "
        f"snippet_chars={sum(len(s) for s in unique_snips)}"
    )
    return expanded, anchor_hits


def load_queries(path="query.txt"):
    queries = []
    for line in open(path, encoding="utf-8"):
        line = re.sub(r"^\d+\.\s*", "", line.strip())
        if line:
            queries.append(line)
    return queries


def ts_to_sec(ts: str) -> float:
    h, m, s = ts.strip().split(":")
    return int(h) * 3600 + int(m) * 60 + float(s)


def clean_screenshot_text(body: str) -> str:
    """Strip OCR layout labels / UI chrome that pollute BM25 keyword recall."""
    text = SCREEN_LABEL_RE.sub("", body)
    text = SCREEN_BOILERPLATE_RE.sub("", text)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text or body.strip()


def load_screenshots(path="doc.txt"):
    raw = open(path, encoding="utf-8").read()
    parts = re.split(r"=+\n时间戳: (\d{2}:\d{2}:\d{2}).*?\n=+\n", raw)
    chunks = []
    for i in range(1, len(parts), 2):
        ts, body = parts[i], parts[i + 1].strip()
        body = re.sub(r"\n=+\n视频处理统计[\s\S]*$", "", body).strip()
        body = clean_screenshot_text(body)
        if body:
            sec = ts_to_sec(ts)
            chunks.append(
                {
                    "timestamp": ts,
                    "start_sec": sec,
                    "end_sec": sec,
                    "text": body,
                    "type": "screen_shot",
                }
            )
    return chunks


def load_transcript(path="transcript.vtt"):
    try:
        raw = open(path, encoding="utf-8").read().strip()
    except FileNotFoundError:
        return []
    if not raw:
        return []
    raw = re.sub(r"^WEBVTT\s*", "", raw, flags=re.I)
    cues = []
    for m in CUE_RE.finditer(raw):
        text = re.sub(r"\s+", " ", m.group(3)).strip()
        if text:
            cues.append({"start": m.group(1), "end": m.group(2), "text": text})

    chunks, buf, start, end, n = [], [], None, None, 0

    def flush():
        nonlocal buf, start, end, n
        if buf:
            chunks.append(
                {
                    "timestamp": f"{start} --> {end}",
                    "start_sec": ts_to_sec(start),
                    "end_sec": ts_to_sec(end),
                    "text": " ".join(buf),
                    "type": "transcript",
                }
            )
        buf, start, end, n = [], None, None, 0

    for c in cues:
        gap = ts_to_sec(c["start"]) - ts_to_sec(end) if end else 0
        if buf and (
            gap > TRANSCRIPT_MAX_GAP_SEC
            or n + len(c["text"]) + 1 > TRANSCRIPT_MAX_CHARS
        ):
            flush()
        if start is None:
            start = c["start"]
        buf.append(c["text"])
        end = c["end"]
        n += len(c["text"]) + 1
    flush()
    return chunks


def timestamp_constraint_value(constraints):
    for c in constraints:
        if c.get("field") == "timestamp":
            return c.get("value", TIMESTAMP)
    return None


def time_distance(center: float, payload: dict) -> float:
    start = payload.get("start_sec")
    end = payload.get("end_sec")
    if start is not None and end is not None:
        if start <= center <= end:
            return 0.0
        return min(abs(start - center), abs(end - center))
    ts = payload.get("timestamp", "")
    if ts and "-->" not in ts:
        try:
            return abs(ts_to_sec(ts) - center)
        except (ValueError, AttributeError):
            pass
    return float("inf")


def _scroll_time_filter(center: float, ts_value: str, doc_type: str | None, *, use_range: bool):
    must = list(BASE_MUST)
    if doc_type:
        must.append(FieldCondition(key="type", match=MatchValue(value=doc_type)))
    if use_range:
        must.extend(
            [
                FieldCondition(
                    key="start_sec",
                    range=Range(lte=center + TIME_WINDOW_SEC),
                ),
                FieldCondition(
                    key="end_sec",
                    range=Range(gte=center - TIME_WINDOW_SEC),
                ),
            ]
        )
    else:
        must.append(
            FieldCondition(key="timestamp", match=MatchValue(value=ts_value))
        )
    return Filter(must=must)


def _scroll_all(client, scroll_filter: Filter, page_size: int = 256):
    """Paginate Qdrant scroll so large time windows are not truncated."""
    out = []
    offset = None
    while True:
        points, offset = client.scroll(
            collection_name="docs",
            scroll_filter=scroll_filter,
            limit=page_size,
            offset=offset,
            with_payload=True,
        )
        out.extend(points)
        if offset is None or not points:
            break
    return out


def search_time_window(client, constraints, doc_type: str | None = None, limit=TIME_NEAR_TOP_K):
    """Filter-only recall: chunks nearest to the playback timestamp (no semantic search)."""
    ts_value = timestamp_constraint_value(constraints)
    if ts_value is None:
        return []

    center = ts_to_sec(ts_value)
    points = _scroll_all(
        client,
        _scroll_time_filter(center, ts_value, doc_type, use_range=True),
    )
    if not points:
        # 兼容旧索引：仅有 timestamp 精确字段、无 start_sec/end_sec
        points = _scroll_all(
            client,
            _scroll_time_filter(center, ts_value, doc_type, use_range=False),
        )

    ranked = sorted(points, key=lambda p: time_distance(center, p.payload))
    kept = []
    for p in ranked[:limit]:
        payload = dict(p.payload)
        payload["time_distance"] = time_distance(center, payload)
        kept.append(SimpleNamespace(id=p.id, payload=payload, score=None))
    return kept



def build_filter(doc_type: str, constraints) -> Filter:
    must = BASE_MUST + [FieldCondition(key="type", match=MatchValue(value=doc_type))]
    ts_value = timestamp_constraint_value(constraints)
    if ts_value is not None:
        # 时间范围重叠：chunk.start <= center+W AND chunk.end >= center-W
        # transcript 存的是区间，不能再用 timestamp KEYWORD eq
        center = ts_to_sec(ts_value)
        must.append(
            FieldCondition(
                key="start_sec",
                range=Range(lte=center + TIME_WINDOW_SEC),
            )
        )
        must.append(
            FieldCondition(
                key="end_sec",
                range=Range(gte=center - TIME_WINDOW_SEC),
            )
        )
    return Filter(must=must)


def search_dense(client, q, doc_type, constraints, limit=DENSE_LIMIT):
    flt = build_filter(doc_type, constraints)
    return client.query_points(
        collection_name="docs",
        query=embed([q])[0],
        using="dense",
        query_filter=flt,
        limit=limit,
    ).points


def search_bm25(client, q, doc_type, constraints, limit=BM25_LIMIT):
    flt = build_filter(doc_type, constraints)
    return client.query_points(
        collection_name="docs",
        query=Document(text=q, model=BM25_MODEL),
        using="bm25",
        query_filter=flt,
        limit=limit,
    ).points


def merge_unique_hits(*hit_lists):
    seen = set()
    merged = []
    for hits in hit_lists:
        for h in hits:
            if h.id in seen:
                continue
            seen.add(h.id)
            merged.append(h)
    return merged


def pick_by_type_quota(hits, top_k=RERANK_TOP_K):
    """Keep both screen_shot and transcript when available (half/half, fill remainder)."""
    if not hits or top_k <= 0:
        return []
    ss = [h for h in hits if h.payload.get("type") == "screen_shot"]
    tr = [h for h in hits if h.payload.get("type") == "transcript"]
    other = [h for h in hits if h.payload.get("type") not in ("screen_shot", "transcript")]
    ss_n = min(len(ss), (top_k + 1) // 2)
    tr_n = min(len(tr), top_k - ss_n)
    # 一侧不足时把名额补给另一侧
    if ss_n + tr_n < top_k:
        ss_n = min(len(ss), top_k - tr_n)
    if ss_n + tr_n < top_k:
        tr_n = min(len(tr), top_k - ss_n)
    picked = merge_unique_hits(ss[:ss_n], tr[:tr_n], other)
    return picked[:top_k]


def rerank_and_filter(query: str, hits, top_k=RERANK_TOP_K, min_score=RERANK_MIN_SCORE):
    if not hits:
        return []
    reranker = get_reranker()
    pairs = [(query, h.payload["text"]) for h in hits]
    scores = reranker.predict(pairs)
    if isinstance(scores, (int, float)):
        scores = [scores]
    ranked = sorted(zip(hits, scores), key=lambda x: x[1], reverse=True)
    scored = []
    for h, s in ranked:
        if s < min_score:
            continue
        payload = dict(h.payload)
        payload["rerank_score"] = float(s)
        scored.append(SimpleNamespace(id=h.id, payload=payload, score=float(s)))
    # 按分数已排序；配额保证 transcript 不被 screen_shot 全部挤掉
    return pick_by_type_quota(scored, top_k)


def build_prompt(question, hits):
    parts = [
        f"Student question: {question}",
        "Lecture content (slides and/or transcript):",
    ]
    if not hits:
        parts.append("(nothing found)")
    else:
        for h in hits:
            p = h.payload
            label = "slide" if p["type"] == "screen_shot" else "transcript"
            parts.append(f"[{label}] ({p['timestamp']})\n{p['text']}")
    return "\n\n".join(parts)


def soften_stiff_refusal(text: str) -> str:
    """Replace machine-style refusals with a natural assistant reply."""
    if STIFF_REFUSAL_RE.match(text.strip()):
        return STIFF_REFUSAL_REPLY
    # 句中仍夹带那句机器话术时，整句换成助手语气
    bad = (
        "i don't know based on the retrieved lecture materials",
        "based on the retrieved lecture materials",
    )
    low = text.lower()
    if any(p in low for p in bad) and ("don't know" in low or "do not know" in low):
        return STIFF_REFUSAL_REPLY
    return text


def answer(prompt: str, *, system: str = ANSWER_SYSTEM) -> str:
    raw = ollama_chat(
        [
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ]
    )
    return soften_stiff_refusal(raw)


def format_hits(hits, channel: str = "") -> str:
    if not hits:
        return ""
    blocks = []
    for h in hits:
        p = h.payload
        score = f"{h.score:.4f}" if h.score is not None else ""
        tag = f" channel={channel}" if channel else ""
        rerank = p.get("rerank_score")
        extra = f" rerank={rerank:.4f}" if rerank is not None else ""
        td = p.get("time_distance")
        if td is not None:
            extra += f" dt={td:.1f}s"
        blocks.append(f"[{p['timestamp']}] score={score}{tag}{extra}\n{p['text']}")
    return "\n\n---\n\n".join(blocks)



ANSWER_HEADERS = [
    "序号",
    "原问题",
    "rewritten_query",
    "hard_constraints",
    "course_id",
    "quarter",
    "lecturer",
    "screen_shot语义召回",
    "screen_shot关键词召回",
    "transcript语义召回",
    "transcript关键词召回",
    "rerank后检索",
    "答案",
]
ANSWER_COL_WIDTHS = {
    "A": 6,
    "B": 36,
    "C": 40,
    "D": 28,
    "E": 12,
    "F": 14,
    "G": 20,
    "H": 42,
    "I": 42,
    "J": 42,
    "K": 42,
    "L": 50,
    "M": 50,
}
CELL_WRAP = Alignment(vertical="top", wrap_text=True)


def init_answer_workbook(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "answers"
    ws.append(ANSWER_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = CELL_WRAP
    for col, width in ANSWER_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    wb.save(path)
    return wb, ws


def append_answer_row(wb, ws, path: str, row):
    ws.append(row)
    for cell in ws[ws.max_row]:
        cell.alignment = CELL_WRAP
    wb.save(path)


parser = argparse.ArgumentParser()
parser.add_argument(
    "--ingest",
    action="store_true",
    help="chunk + embed + upsert; default skips ingest and searches existing data",
)
parser.add_argument(
    "--backfill-bm25-phrases",
    action="store_true",
    help="rebuild BM25 sparse vectors with phrase tokens from existing payloads (no dense re-embed)",
)
parser.add_argument(
    "--inspect-phrase",
    type=str,
    default=None,
    help='compare bag BM25 vs phrase-token BM25, e.g. --inspect-phrase "Question 9"',
)
args = parser.parse_args()

client = QdrantClient(url="http://localhost:6333", check_compatibility=False)

if args.inspect_phrase:
    phrase = normalize_anchor_phrase(args.inspect_phrase)
    token = phrase_token(phrase)
    print(f"inspect phrase: {phrase} → token {token}")
    bag = search_bm25(client, phrase, "screen_shot", [], limit=5)
    print(f"\nBM25 bag query {phrase!r}:")
    for h in bag:
        t = h.payload.get("text", "")
        print(
            f"  id={h.id} score={h.score:.4f} "
            f"literal={text_has_phrase(t, phrase)} ts={h.payload.get('timestamp')}"
        )
        print("   ", t[:120].replace("\n", " / "))
    ph = search_phrase_bm25(client, phrase, "screen_shot", [], limit=5)
    print(f"\nBM25 phrase-token query {token!r}:")
    for h in ph:
        t = h.payload.get("text", "")
        print(
            f"  id={h.id} score={h.score:.4f} "
            f"literal={text_has_phrase(t, phrase)} ts={h.payload.get('timestamp')}"
        )
        print("   ", t[:120].replace("\n", " / "))
    raise SystemExit(0)

if args.backfill_bm25_phrases:
    # 只重写 bm25 稀疏向量，不重算 dense
    offset = None
    updated = with_tok = 0
    while True:
        points, offset = client.scroll(
            collection_name="docs",
            limit=64,
            offset=offset,
            with_payload=True,
        )
        if not points:
            break
        batch_vecs = []
        for p in points:
            text = p.payload.get("text", "")
            idx = bm25_index_text(text)
            if idx != text:
                with_tok += 1
            batch_vecs.append(
                PointVectors(
                    id=p.id,
                    vector={"bm25": Document(text=idx, model=BM25_MODEL)},
                )
            )
            updated += 1
        client.update_vectors(collection_name="docs", points=batch_vecs)
        print(f"backfill bm25 phrases {updated} (with tokens {with_tok})")
        if offset is None:
            break
    print(f"done: updated={updated} with_phrase_tokens={with_tok}")
    raise SystemExit(0)

if args.ingest:
    chunks = load_screenshots() + load_transcript()
    print(
        f"screen_shot {sum(c['type']=='screen_shot' for c in chunks)} 段, "
        f"transcript {sum(c['type']=='transcript' for c in chunks)} 段"
    )

    vectors = embed([c["text"] for c in chunks])

    if client.collection_exists("docs"):
        client.delete_collection("docs")

    client.create_collection(
        "docs",
        vectors_config={
            "dense": VectorParams(size=len(vectors[0]), distance=Distance.COSINE),
        },
        sparse_vectors_config={
            "bm25": SparseVectorParams(modifier=Modifier.IDF),
        },
    )
    for field in ("course_id", "quarter", "lecturer", "type", "timestamp"):
        client.create_payload_index(
            "docs", field, field_schema=PayloadSchemaType.KEYWORD
        )
    for field in ("start_sec", "end_sec"):
        client.create_payload_index(
            "docs", field, field_schema=PayloadSchemaType.FLOAT
        )

    UPSERT_BATCH = 64
    total = len(chunks)
    for start in range(0, total, UPSERT_BATCH):
        batch = chunks[start : start + UPSERT_BATCH]
        client.upsert(
            "docs",
            points=[
                PointStruct(
                    id=start + j,
                    vector={
                        "dense": vectors[start + j],
                        # BM25 用带短语整体 token 的文本；payload.text 仍是干净原文
                        "bm25": Document(
                            text=bm25_index_text(c["text"]), model=BM25_MODEL
                        ),
                    },
                    payload={
                        "text": c["text"],
                        "timestamp": c["timestamp"],
                        "start_sec": c["start_sec"],
                        "end_sec": c["end_sec"],
                        "type": c["type"],
                        "course_id": COURSE_ID,
                        "quarter": QUARTER,
                        "lecturer": LECTURER,
                    },
                )
                for j, c in enumerate(batch)
            ],
        )
        print(f"upsert {min(start + UPSERT_BATCH, total)}/{total}")
else:
    print("skip ingest, search existing collection")

stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
out_path = f"answer_{stamp}.xlsx"
wb, ws = init_answer_workbook(out_path)
print(f"writing answers incrementally to {out_path}")

for i, query in enumerate(load_queries(), 1):
    rewritten = rewrite(query)
    q = rewritten["rewritten_query"]
    constraints = rewritten["hard_constraints"]
    print(f"\nquery:    {query}")
    print("rewrite:")
    print(json.dumps(rewritten, ensure_ascii=False, indent=2))

    ts_value = timestamp_constraint_value(constraints)

    if ts_value:
        # “现在在讲什么”：只按时间邻近取内容，不做语义检索 / rerank
        # 每类各取 top_k，再按类型配额合并，避免截图挤掉 transcript
        tw_ss = search_time_window(
            client, constraints, "screen_shot", limit=TIME_NEAR_TOP_K
        )
        tw_tr = search_time_window(
            client, constraints, "transcript", limit=TIME_NEAR_TOP_K
        )
        center = ts_to_sec(ts_value)
        # 各类内部已按时间距离排序；配额合并保证两边都进最终结果
        final_hits = pick_by_type_quota(
            merge_unique_hits(tw_ss, tw_tr),
            TIME_NEAR_TOP_K,
        )
        # 最终再按时间距离排一下，方便阅读
        final_hits = sorted(
            final_hits,
            key=lambda h: time_distance(center, h.payload),
        )
        ss_dense = tw_ss
        ss_bm25 = []
        tr_dense = tw_tr
        tr_bm25 = []
        print(
            f"time-only recall ss={len(tw_ss)} tr={len(tw_tr)} → kept {len(final_hits)} "
            f"(ss={sum(1 for h in final_hits if h.payload.get('type')=='screen_shot')} "
            f"tr={sum(1 for h in final_hits if h.payload.get('type')=='transcript')})"
        )
        if not final_hits:
            print(
                "warning: timestamp filter matched 0 chunks — "
                "re-run with --ingest to rebuild start_sec/end_sec indexes"
            )
            prompt = build_prompt(query, final_hits)
            answer_text = NO_HIT_NOW_REPLY
        else:
            prompt = build_prompt(query, final_hits)
            answer_text = answer(prompt, system=NOW_ANSWER_SYSTEM)
    else:
        # 题号类问题：先锚定找回题干，再扩充 query（dense 仍用短 query，避免向量被长 OCR 冲淡）
        q_short = q
        q_expand, anchor_hits = expand_query_with_anchors(
            client, query, q_short, constraints
        )
        q = q_expand  # excel / 日志里展示扩充后的 rewritten
        if anchor_hits:
            print("rewrite (after anchor expand):")
            print(q[:500] + ("..." if len(q) > 500 else ""))

        ss_dense = search_dense(client, q_short, "screen_shot", constraints)
        ss_bm25 = search_bm25(client, q_expand, "screen_shot", constraints)
        tr_dense = search_dense(client, q_short, "transcript", constraints)
        tr_bm25 = search_bm25(client, q_expand, "transcript", constraints)
        print(
            f"recall screen_shot dense={len(ss_dense)} bm25={len(ss_bm25)} | "
            f"transcript dense={len(tr_dense)} bm25={len(tr_bm25)} | "
            f"anchors={len(anchor_hits)}"
        )
        candidates = merge_unique_hits(
            anchor_hits, ss_dense, ss_bm25, tr_dense, tr_bm25
        )
        rerank_query = f"{query}\n{q_expand}"
        final_hits = rerank_and_filter(rerank_query, candidates)
        print(f"rerank kept {len(final_hits)}/{len(candidates)}")
        prompt = build_prompt(query, final_hits)
        if not final_hits:
            answer_text = NO_HIT_REPLY
        else:
            answer_text = answer(prompt)

    print(prompt)
    print(f"\nanswer:\n{answer_text}")

    append_answer_row(
        wb,
        ws,
        out_path,
        [
            i,
            query,
            q,
            json.dumps(constraints, ensure_ascii=False),
            COURSE_ID,
            QUARTER,
            LECTURER,
            format_hits(ss_dense, "semantic" if not ts_value else "time"),
            format_hits(ss_bm25, "keyword"),
            format_hits(tr_dense, "semantic" if not ts_value else "time"),
            format_hits(tr_bm25, "keyword"),
            format_hits(final_hits, "time" if ts_value else "rerank"),
            answer_text,
        ],
    )
    print(f"wrote row {i} → {out_path}")

print(f"done: {out_path}")
